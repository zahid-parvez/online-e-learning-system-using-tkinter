import smtplib
import sqlite3

from tkinter import *

from tkinter import messagebox

from openpyxl import load_workbook, Workbook

import password_generator


def new_user():
    new_user_sign_up_window = Tk()
    new_user_sign_up_window.title("  Welcome  ")
    new_user_sign_up_window.geometry("900x500")
    Label(new_user_sign_up_window, text="   ** USER's SIGN-UP **   ", fg='indigo', font=("Wide Latin", 27),
          bg="grey").grid(row=0, column=0)
    user_name_label = Label(new_user_sign_up_window, text="Enter your name : ", padx=2, pady=2,
                            font=('Arial Rounded MT Bold', 12))
    user_name_label.grid(row=1, column=0)
    user_name_entry = Entry(new_user_sign_up_window, width=20, borderwidth=5)
    user_name_entry.grid(row=1, column=1)

    user_age_label = Label(new_user_sign_up_window, text="Enter your age : ", padx=2, pady=2,
                           font=('Arial Rounded MT Bold', 12))
    user_age_label.grid(row=2, column=0)
    user_age_entry = Entry(new_user_sign_up_window, width=20, borderwidth=5)
    user_age_entry.grid(row=2, column=1)

    user_DOB_label = Label(new_user_sign_up_window, text="Select your Date of Birth : ", padx=2, pady=2,
                           font=('Arial Rounded MT Bold', 12))
    user_DOB_label.grid(row=3, column=0)
    user_DOB_entry = Entry(new_user_sign_up_window, width=20, borderwidth=5)
    user_DOB_entry.grid(row=3, column=1)

    user_gender_label = Label(new_user_sign_up_window, text="Select your gender : ", font=('Arial Rounded MT Bold', 12))
    user_gender_label.grid(row=4, column=0)
    user_gender_entry = Entry(new_user_sign_up_window, width=20, borderwidth=5)
    user_gender_entry.grid(row=4, column=1)

    user_college_name_label = Label(new_user_sign_up_window, text="Select your college name : ", padx=2, pady=2,
                                    font=('Arial Rounded MT Bold', 12))
    user_college_name_label.grid(row=5, column=0)
    user_college_name_entry = Entry(new_user_sign_up_window, width=20, borderwidth=5)
    user_college_name_entry.grid(row=5, column=1)

    user_college_course_name_label = Label(new_user_sign_up_window, text="Select your college course name : ", padx=2,
                                           pady=2, font=('Arial Rounded MT Bold', 12))
    user_college_course_name_label.grid(row=6, column=0)
    user_college_course_name_entry = Entry(new_user_sign_up_window, width=20, borderwidth=5)
    user_college_course_name_entry.grid(row=6, column=1)

    user_college_course_year_label = Label(new_user_sign_up_window, text="Enter the year in which you are studying : ",
                                           font=('Arial Rounded MT Bold', 12), padx=2, pady=2)
    user_college_course_year_label.grid(row=7, column=0)
    user_college_course_year_entry = Entry(new_user_sign_up_window, width=20, borderwidth=5)
    user_college_course_year_entry.grid(row=7, column=1)

    user_college_course_dept_label = Label(new_user_sign_up_window, text="Select your dept : ",
                                           padx=2, pady=2, font=('Arial Rounded MT Bold', 12))
    user_college_course_dept_label.grid(row=8, column=0)
    user_college_course_dept_entry = Entry(new_user_sign_up_window, width=20, borderwidth=5)
    user_college_course_dept_entry.grid(row=8, column=1)

    user_phone_number_label = Label(new_user_sign_up_window, text="Enter your mobile number : ",
                                    font=('Arial Rounded MT Bold', 12))
    user_phone_number_label.grid(row=9, column=0)
    user_phone_number_entry = Entry(new_user_sign_up_window, width=20, borderwidth=5)
    user_phone_number_entry.grid(row=9, column=1)

    user_email_id_label = Label(new_user_sign_up_window, text="Enter your email-id : ",
                                font=('Arial Rounded MT Bold', 12))
    user_email_id_label.grid(row=10, column=0)
    user_email_id_entry = Entry(new_user_sign_up_window, width=20, borderwidth=5)
    user_email_id_entry.grid(row=10, column=1)

    def save_details():
        user_college_course_dept = user_college_course_dept_entry.get()
        user_college_course_year = user_college_course_year_entry.get()
        user_college_course_name = user_college_course_name_entry.get()
        user_college_name = user_college_name_entry.get()
        user_gender = user_gender_entry.get()
        user_DOB = user_DOB_entry.get()
        user_age = user_age_entry.get()
        user_name = user_name_entry.get()
        user_phone_number = user_phone_number_entry.get()
        user_email_id = user_email_id_entry.get()

        detail_list = [user_name,
                       user_age,
                       user_DOB,
                       user_gender,
                       user_phone_number,
                       user_email_id,
                       user_college_name,
                       user_college_course_year,
                       user_college_course_name,
                       user_college_course_dept
                       ]

        # with open('dataSource.json', 'a') as json_file:
        #     js.dump(details, json_file)
        #     json_file.write("\n")
        password = password_generator.get_password()
        # create a workbook object
        wb = Workbook()
        # load existing sheet
        wb = load_workbook('details.xlsx')
        # create an active worksheet
        ws = wb.active
        # row1 = ws["2"]
        # for i in row1:
        #     print(i.value)
        list_r = {'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'}
        # list_c = {'1', '2', '3', '4', '5', '6', '7', '8', '9'}
        _col = 2
        for row, detail in zip(list_r, detail_list):
            a = row + str(_col)
            ws[a] = detail
        wb.save('details.xlsx')
        '''
        conn = sqlite3.connect('student_details_database.db')

        c = conn.cursor()

        c.execute(""" CREATE TABLE addresses (
                       user_name text,
                       user_age text,
                       user_DOB text,
                       user_gender text,
                       user_phone_number text,
                       user_email_id text,
                       user_college_name text,
                       user_college_course_year text,
                       user_college_course_name text,
                       user_college_course_dept text
                       )""")

        c.execute(""" INSERT INTO addresses VALUES (:user_name,:user_age,:user_DOB,:user_gender,:user_phone_number,
        user_email_id,:user_college_name,:user_college_course_year,:user_college_course_name,:user_college_course_dept)",)
                    {
                            'user_name' : user_n:me_entry.get(),
                            'user_age' : user_age_entry.get(),
                            'user_DOB' : user_DOB_entry.get(),
                            'user_gender' : user_gender_entry.get(),    
                            'user_phone_number' : user_phone_number_entry.get(),
                            'user_email_id' : user_email_id_entry.get(),
                            'user_college_name' : user_college_name_entry.get()
                            'user_college_course_year' : user_college_course_year_entry.get()
                            'user_college_course_name' : user_college_course_name_entry.get()
                             'user_college_course_dept' : user_college_course_dept_entry.get()
                            }
                            """)
        conn.commit()
        conn.close()
        '''
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
        msg1 = \
            f'Hello {user_name},  Welcome to our learning platform. ' \
            f'Your password is vasavi@1 and user id is {user_name}. ' \
            f'Now you can login into our website.'
        server.login("tminiproject@gmail.com", "Mnpr@t456")
        server.sendmail("miniproject@gmail.com", user_email_id, msg1)
        server.quit()
        messagebox.showinfo(title=f"Welcome the learning world {user_name}",
                            message="Please check your mail for login details")

    Button(new_user_sign_up_window, text="     ** Save **     ", bg="Blue", activeforeground="Blue",
           font=('Arial Rounded MT Bold', 12), command=save_details).grid(row=12, column=0)
    Button(new_user_sign_up_window, text="   Exit  ", bg='red', fg='cyan',
           font=('Arial Rounded MT Bold', 12), command=new_user_sign_up_window.destroy).grid(row=12, column=2)
    new_user_sign_up_window.mainloop()


# new_user()
