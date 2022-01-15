from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


def view_student_details():
    # create a workbook object
    wb = Workbook()
    # load existing sheet
    wb = load_workbook('details.xlsx')
    # create an active worksheet
    ws = wb.active
    students_details_window = Tk()
    students_details_window.title("student details")
    students_details_window.geometry("1250x700")
    Label(students_details_window, text="   **student details**   ", fg='indigo', font=("Wide Latin", 27),
          bg="grey").pack()
    col_A = ws['A']
    col_B = ws['B']
    col_C = ws['C']
    col_D = ws['D']
    col_E = ws['E']
    col_F = ws['F']
    col_G = ws['G']
    col_H = ws['H']
    col_I = ws['I']
    col_J = ws['J']

    def fun():
        list1 = ''
        for cell in col_A:
            list1 = f'{list1 + str(cell.value)}\n'

        label_a.config(text=list1)

        list2 = ''
        for cell in col_B:
            list2 = f'{list2 + str(cell.value)}\n'

        label_b.config(text=list2)

        list3 = ''
        for cell in col_C:
            list3 = f'{list3 + str(cell.value)}\n'

        label_c.config(text=list3)

        list4 = ''
        for cell in col_D:
            list4 = f'{list4 + str(cell.value)}\n'

        label_d.config(text=list4)

        list5 = ''
        for cell in col_E:
            list5 = f'{list5 + str(cell.value)}\n'

        label_e.config(text=list5)

        list6 = ''
        for cell in col_F:
            list6 = f'{list6 + str(cell.value)}\n'

        label_f.config(text=list6)

        list7 = ''
        for cell in col_G:
            list7 = f'{list7 + str(cell.value)}\n'

        label_g.config(text=list7)

        list8 = ''
        for cell in col_H:
            list8 = f'{list8 + str(cell.value)}\n'

        label_h.config(text=list8)

        list9 = ''
        for cell in col_I:
            list9 = f'{list9 + str(cell.value)}\n'

        label_i.config(text=list9)

        list10 = ''
        for cell in col_J:
            list10 = f'{list10 + str(cell.value)}\n'

        label_j.config(text=list10)

    Button(students_details_window, text='   Get Details of students registered in courses  ', command=fun,
           font=('Arial Rounded MT Bold', 12), bg='blue', fg='light blue',
           activeforeground='green').pack(anchor=CENTER, padx=5, pady=10)

    Button(students_details_window, text='   Exit   ', command=students_details_window.destroy,
           font=('Arial Rounded MT Bold', 12), bg='purple', fg='violet', activeforeground='indigo').place(x=600, y=500)

    label_a = Label(students_details_window, text='')
    label_a.place(x=0, y=100)
    label_b = Label(students_details_window, text='')
    label_b.place(x=140, y=100)
    label_c = Label(students_details_window, text='')
    label_c.place(x=260, y=100)
    label_d = Label(students_details_window, text='')
    label_d.place(x=380, y=100)
    label_e = Label(students_details_window, text='')
    label_e.place(x=460, y=100)
    label_f = Label(students_details_window, text='')
    label_f.place(x=580, y=100)
    label_g = Label(students_details_window, text='')
    label_g.place(x=800, y=100)
    label_h = Label(students_details_window, text='')
    label_h.place(x=1025, y=100)
    label_i = Label(students_details_window, text='')
    label_i.place(x=1125, y=100)
    label_j = Label(students_details_window, text='')
    label_j.place(x=1200, y=100)
    students_details_window.mainloop()


# view_student_details()
